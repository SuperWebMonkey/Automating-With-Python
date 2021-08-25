#  https://www.youtube.com/watch?v=7YS6YDQKFh0 , on 11:00
#  python for data analysis: https://www.youtube.com/watch?v=r-uOLxNrNk8 
#  -- currently on 14:30

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def menu():
    print('1) Add data\n','2) Delete data\n','0) quit the program\n')

def option():
    menu()
    print('Please select one of the following options above:')
    option =  input()

    while True:
        if option == 0:
            print('You are now quiting the programing')
            break
        elif option == 1:
            print('You are adding to the excel sheeting')
        elif option == 2:
            print('You are deleting from the excel sheet')

menu()

wb = load_workbook('excel-files/jobs.xlsx')
ws = wb.active

print(wb.sheetnames)
